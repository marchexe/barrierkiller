import os
from openpyxl import load_workbook
from google.cloud import texttospeech
from moviepy import TextClip, concatenate_videoclips, AudioFileClip, ColorClip, vfx, afx

os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "service_account.json"
client = texttospeech.TextToSpeechClient()

WIDTH, HEIGHT = 1920, 1080
FONT = "dejavu-sans-book.otf"
FONT_SIZE = 60
TEXT_COLOR = (0, 0, 0)
BG_COLOR = (255, 255, 255)

workbook = load_workbook("vocab.xlsx")
sheet = workbook.active

os.makedirs("components", exist_ok=True)
os.makedirs("output", exist_ok=True)
os.makedirs("video_output", exist_ok=True)

max_rows = 3

columns_order = ["de", "ru", "b1_de", "b1_ru", "b2_de", "b2_ru"]

voice_map = {
    "de": texttospeech.VoiceSelectionParams(language_code="de-DE", name="de-DE-Studio-B"),
    "ru": texttospeech.VoiceSelectionParams(language_code="ru-RU", name="ru-RU-Wavenet-D"),
    "b1_de": texttospeech.VoiceSelectionParams(language_code="de-DE", name="de-DE-Studio-B"),
    "b1_ru": texttospeech.VoiceSelectionParams(language_code="ru-RU", name="ru-RU-Wavenet-D"),
    "b2_de": texttospeech.VoiceSelectionParams(language_code="de-DE", name="de-DE-Studio-C"),
    "b2_ru": texttospeech.VoiceSelectionParams(language_code="ru-RU", name="ru-RU-Wavenet-C")
}

audio_config = texttospeech.AudioConfig(
    audio_encoding=texttospeech.AudioEncoding.MP3,
    speaking_rate=1.0,
    pitch=1
)


def generate_speech(text, voice, filename):
    if not text or not text.strip():
        return None

    path = f"components/{filename}"
    if os.path.exists(path):
        return path

    synthesis_input = texttospeech.SynthesisInput(text=text)
    response = client.synthesize_speech(
        input=synthesis_input, voice=voice, audio_config=audio_config
    )

    with open(path, "wb") as f:
        f.write(response.audio_content)

    return path if os.path.exists(path) else None


row_video_clips = []

for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    if row_idx > max_rows:
        break

    row_clips = []

    for col_idx, text in enumerate(row):
        if col_idx >= len(columns_order):
            break

        col_name = columns_order[col_idx]
        voice = voice_map.get(col_name)
        filename = f"{col_name}_{row_idx}.mp3"
        audio_path = generate_speech(text, voice, filename)

        if text and text.strip():
            duration = 1
            if audio_path:
                audio_clip = AudioFileClip(audio_path)
                duration = max(audio_clip.duration, 1)

            intra_row_pause = 1
            fade_duration = intra_row_pause / 2

            text_clip = TextClip(
                text=text,
                font_size=FONT_SIZE,
                color=TEXT_COLOR,
                font=FONT,
                size=(WIDTH, HEIGHT),
                bg_color=BG_COLOR,
                method="caption",
                text_align="center"
            ).with_duration(duration)

            if audio_path:
                text_clip = text_clip.with_audio(audio_clip)

            text_clip = text_clip.with_effects([
                vfx.CrossFadeIn(fade_duration),
                vfx.CrossFadeOut(fade_duration),
                afx.AudioFadeIn(fade_duration),
                afx.AudioFadeOut(fade_duration)
            ])

            row_clips.append(text_clip)

            pause_clip = ColorClip(size=(WIDTH, HEIGHT),
                                   color=BG_COLOR, duration=intra_row_pause)
            row_clips.append(pause_clip)

    if row_clips:
        row_clips.pop()

    row_video_clips.extend(row_clips)

    inter_row_pause = 2
    fade_duration = inter_row_pause / 2

    if row_video_clips:
        row_video_clips.append(
            ColorClip(size=(WIDTH, HEIGHT), color=BG_COLOR, duration=inter_row_pause))

if row_video_clips:
    row_video_clips.pop()

if row_video_clips:
    final_video = concatenate_videoclips(
        row_video_clips, method="compose", bg_color=BG_COLOR)
    final_video.write_videofile(
        "video_output/final_video.mp4", fps=24, codec="libx264")
    print("Video successfully created: video_output/final_video.mp4")
else:
    print("No text for video making was found.")
