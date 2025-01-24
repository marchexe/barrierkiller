import os
from pydub import AudioSegment
from openpyxl import load_workbook
from moviepy import TextClip, CompositeVideoClip, ColorClip, AudioFileClip

width, height = 1920, 1080
font_size = 80
font_path="/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
text_color = (0, 0, 0)
background_color = (255, 255, 255)
fade_duration = 0.5

wb = load_workbook("vocab.xlsx")
ws = wb.active

audio_clip = AudioFileClip("output/final.mp3")

background = ColorClip((width, height), color=background_color,
                      duration=audio_clip.duration)

columns_order = ["de", "ru", "b1_de", "b1_ru",
                 "b2_de", "b2_ru", "c1_de", "c1_ru"]

subtitles = []
current_time = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    for col_idx, text in enumerate(row):
        if col_idx >= len(columns_order):
            break

        if text:
            segment = AudioSegment.from_file(
                f"components/{columns_order[col_idx]}_{row_idx}.mp3")
            duration = len(segment) / 1000

            txt_clip = TextClip(
                text, font_size=font_size, color=text_color, font=font_path, size=(width, None), method="caption"
            ).set_position("center").set_duration(duration).fadein(fade_duration).fadeout(fade_duration)

            txt_clip = txt_clip.set_start(current_time)
            subtitles.append(txt_clip)

            current_time += duration + \
                (1 if col_idx < len(columns_order) - 1 else 2)

final_video = CompositeVideoClip([background] + subtitles).set_audio(audio_clip)
final_video.write_videofile(
    "output/final.mp4", fps=30, codec="libx264", audio_codec="aac")
