import os
from pydub import AudioSegment
from openpyxl import load_workbook
from google.cloud import texttospeech

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'service_account.json'
client = texttospeech.TextToSpeechClient()

wb = load_workbook("vocab.xlsx")
ws = wb.active

os.makedirs('components', exist_ok=True)
os.makedirs('output', exist_ok=True)

max_rows = 2

columns_order = ["de", "ru", "b1_de", "b1_ru", "b2_de", "b2_ru"]

voice_map = {
    "de": texttospeech.VoiceSelectionParams(language_code='de-DE', name='de-DE-Studio-B'),
    "ru": texttospeech.VoiceSelectionParams(language_code='ru-RU', name='ru-RU-Wavenet-D'),
    "b2_de": texttospeech.VoiceSelectionParams(language_code='de-DE', name='de-DE-Studio-C'),
    "b2_ru": texttospeech.VoiceSelectionParams(language_code='ru-RU', name='ru-RU-Wavenet-C')
}

audio_config = texttospeech.AudioConfig(
    audio_encoding=texttospeech.AudioEncoding.MP3,
    speaking_rate=1.0,
    pitch=1
)


def generate_speech(text, voice, filename):
    if not text or not text.strip():
        return None

    path = f"components/{filename}"

    if os.path.exists(path):
        print(f"File already exists: {path}")
        return AudioSegment.from_file(path)

    print(f"Generating a new file: {path}")
    synthesis_input = texttospeech.SynthesisInput(text=text)
    response = client.synthesize_speech(
        input=synthesis_input, voice=voice, audio_config=audio_config)

    with open(path, "wb") as f:
        f.write(response.audio_content)
    return AudioSegment.from_file(path)


segments = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    if row_idx > max_rows:
        break

    row_segments = []
    last_german_audio = None

    for col_idx, text in enumerate(row):
        if col_idx >= len(columns_order):
            break

        col_name = columns_order[col_idx]

        if "b2" in col_name:
            voice_key = "b2_de" if "de" in col_name else "b2_ru"
        else:
            voice_key = "de" if "de" in col_name else "ru"

        voice = voice_map[voice_key]
        filename = f"{col_name}_{row_idx}.mp3"
        audio = generate_speech(text, voice, filename)

        if audio:
            row_segments.append(audio)
            row_segments.append(AudioSegment.silent(duration=1000))

            if "b1_de" in col_name or "b2_de" in col_name:
                last_german_audio = audio

        if ("b1_ru" in col_name or "b2_ru" in col_name) and last_german_audio:
            row_segments.append(last_german_audio)
            row_segments.append(AudioSegment.silent(duration=1000))

    if row_segments:
        row_segments.pop()
        segments.extend(row_segments)
        segments.append(AudioSegment.silent(duration=2000))


if segments:
    final_audio = sum(segments)
    final_audio.export("output/final.mp3", format="mp3")
    print("Audiofile created: output/final.mp3")
